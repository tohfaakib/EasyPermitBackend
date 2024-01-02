import requests
import openpyxl
import os
import json
from pyairtable import Api
from settings import AppSettings

settings = AppSettings()

def get_usgeocoder_data(smartyAutoAddress):
    url = "https://usgeocoder.com/api/get_info.php"
    params = {
        "address": smartyAutoAddress['street_line'], 
        "zipcode": smartyAutoAddress['zipcode'], 
        "authkey": settings.GEOCODER_AUTH_KEY, 
        "format": "json"
    }
    us_geocoder_response = requests.get(url, params=params)
    us_geocoder_data = json.loads(us_geocoder_response.text.lstrip('\ufeff'))
    # print("Usgeocoder_Data: ", us_geocoder_data)
    return us_geocoder_data

def get_melissa_property_data(address):
    url = 'https://property.melissadata.net/v4/WEB/LookupProperty/'
    params = {
        'id': settings.PROPERTY_MELISSADATA_ID,
        'ff': address,
        'format': 'json'
    }
    response = requests.get(url, params=params)
    melisssa_data = json.loads(response.text.lstrip('\ufeff'))
    # print("Melissa_Property_Data: ", melisssa_data)
    return melisssa_data

def get_melissa_address_data(address, city, state):
    url = 'https://address.melissadata.net/v3/WEB/GlobalAddress/doGlobalAddress'

    params = {
        'id': settings.ADDRESS_MELISSADATA_ID,
        'a1': address,
        'loc': city,
        'ctry=USA&admarea=': state,
        'format': 'json'
    }
    response = requests.get(url, params=params)
    address_data = json.loads(response.text.lstrip('\ufeff'))
    # print("Melissa_Address_Data: ", address_data)
    return address_data

def get_corresponding_value(zipcode):
    excel_file_path = os.path.join(os.path.dirname(__file__), 'BuildingClimateZonesByZIPCode_ada.xlsx')
    
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    
    # Assume the data is in the first sheet and starts from the second row
    worksheet = workbook.active
    
    # Iterate through the rows
    for i in range(2, worksheet.max_row + 1):
        zipcode_cell = worksheet.cell(row=i, column=1).value
        value_cell = worksheet.cell(row=i, column=2).value
        
        if not (zipcode_cell and value_cell):
            # No more data
            break
        
        current_zipcode = str(zipcode_cell)
        if current_zipcode == zipcode:
            return value_cell


def property_details(smartyAutoAddress):
    smarty_street = smartyAutoAddress['street_line']
    smarty_city = smartyAutoAddress['city']
    smarty_zipcode = smartyAutoAddress['zipcode']
    smarty_state = smartyAutoAddress['state']
    # print(smarty_street, smarty_city, smarty_state, smarty_zipcode)

    try:
        # us_geocoder_data = get_usgeocoder_data(smartyAutoAddress)
        property_data = get_melissa_property_data(f'{smarty_street}, {smarty_city}, {smarty_state} {smarty_zipcode}')
        # address_data = get_melissa_address_data(f'{smarty_street}, {smarty_city}, {smarty_state} {smarty_zipcode}', smarty_city, smarty_state)
        cliemate_zone = get_corresponding_value(smarty_zipcode)

        data = {
            "climate_zone": cliemate_zone,
            "full_street_address": f'{smarty_street}, {smarty_city}, {smarty_state}, {smarty_zipcode}',
            "unit": 'unit',
            "apn": property_data['Records'][0]['Parcel']['UnformattedAPN'],
            "owner": property_data['Records'][0]['PrimaryOwner']['Name1Full'],
            "year_built": property_data['Records'][0]['PropertyUseInfo']['YearBuilt'],
            "square_feet": property_data['Records'][0]['PropertySize']['AreaBuilding'],
            "lot_size": property_data['Records'][0]['PropertySize']['AreaLotSF'],
            "bedrooms": property_data['Records'][0]['IntRoomInfo']['BedroomsCount'],
            "total_rooms": property_data['Records'][0]['IntRoomInfo']['RoomsCount'],
        }

        print("all:", data)

        return data
    except Exception as e:
        print(e)

def geocoder_municipality_status(smartyAutoAddress, filtered_airtable_cities):
    url = "https://usgeocoder.com/api/get_info.php"
    params = {
        "address": f"{smartyAutoAddress['street_line']}", 
        "zipcode": smartyAutoAddress['zipcode'], 
        "authkey": settings.GEOCODER_AUTH_KEY, 
        "format": "json"
    }
    try:
        response = requests.get(url, params=params)
        data = json.loads(response.text)

        municipal_status = data['usgeocoder']['jurisdictions_info']['municipal']['municipal_status']

        if municipal_status == 'Match Found':
            municipal_name = data['usgeocoder']['jurisdictions_info']['municipal']['municipal_name']

            if municipal_name in filtered_airtable_cities:
                response = {
                    'data': {
                        'message': "Municipality Match Found and municipal_name(city) in airtable_cities",
                        'success': True
                    },
                }
                return response
            else:
                response = {
                    'data': {
                        'message': "Municipality Match Found but municipal_name(city) not in airtable_cities!",
                        'success': False
                    },
                }
                return response
            
        elif municipal_status == 'Un-incorporated':
            if '-' in filtered_airtable_cities:
                response = {
                    'data': {
                        'message': "Municipality Un-incorporated and - in airtable_cities.",
                        'success': True
                    },
                }
                return response
            else:
                response = {
                    'data': {
                        'message': "Municipality Un-incorporated and - not in airtable_cities.",
                        'success': False
                    },
                }
                return response

    except Exception as e:
        response = {
            'data': {
                'message': "Geocoder VPN is turned OFF.",
                'success': False
            },
        }
        return response

def airtable_county_check(smartyAutoAddress, smarty_county):
    api = Api(settings.AIRTABLE_TOKEN)
    table = api.table(settings.AIRTABLE_BASE_ID, settings.AIRTABLE_TABLE_ID)

    filter_formula = f"AND({{Allowed?}}=1, {{county}}='{smarty_county}')"

    # filter_formula = "{Allowed?}=1"
    filtered_airtable = table.all(formula=filter_formula)

    filtered_airtable_cities = [item['fields']['City'] for item in filtered_airtable]

    if len(filtered_airtable) > 0:
        return geocoder_municipality_status(smartyAutoAddress, filtered_airtable_cities)
    else:
        response = {
            'data': {
                'message': "Smarty county is not available in Airtable county.",
                'success': False
            },
        }
        return response

def smarty_residency_county(smartyAutoAddress):
    url = "https://us-street.api.smarty.com/street-address"
    params = {
        "auth-id": settings.SMARTY_AUTH_ID,
        "auth-token": settings.SMARTY_AUTH_TOKEN,
        "street": smartyAutoAddress['street_line'],
        "city": smartyAutoAddress['city'],
        "state": smartyAutoAddress['state']
    }

    response = requests.get(url, params=params)
    
    if response.status_code == 200:
        res = response.json()
        
        try:
            residency = res[0]['metadata']['rdi'] if res else None
            county_name = res[0]['metadata']['county_name'] if res else None

            data = {
                "residency": residency,
                "county": county_name
            }
            return data
        except Exception as e:
            data = {
                "residency": str(e),
                "county": str(e)
            }
            return data
    else:
        print(f"Request failed with status code {response.status_code}")
        return None


def validate_address(smartyAutoAddress):
    if smartyAutoAddress['state'] == 'CA':
        rc = smarty_residency_county(smartyAutoAddress)
        smarty_street = smartyAutoAddress['street_line']
        smarty_city = smartyAutoAddress['city']
        smarty_zipcode = smartyAutoAddress['zipcode']
        smarty_state = smartyAutoAddress['state']
        
        # property_data = get_melissa_property_data(f'{smarty_street}, {smarty_city}, {smarty_state} {smarty_zipcode}')
        # room_count = int(property_data['Records'][0]['IntRoomInfo']['BathCount']) + int(property_data['Records'][0]['IntRoomInfo']['BedroomsCount']) + int(property_data['Records'][0]['IntRoomInfo']['RoomsCount'])
        property_data = property_details(smartyAutoAddress)

        # if room_count != 0:
        #     county_response = airtable_county_check(smartyAutoAddress, rc['county'])
        #     return county_response
        # response = {
        #     'data': {
        #         'message': "Address is  not residential.",
        #         'success': False
        #     },
        # }
        # return response
        return  {
            'data': {
                'data': property_data,
                'message': "Address is  not residential.",
                'success': True
            },
        }
    else:
        response = {
            'data': {
                'message': "Sorry! Service only available for california state",
                'success': False
            },
        }
        return response

from pymongo.mongo_client import MongoClient
from pymongo.errors import ConnectionFailure

# Replace placeholders with actual MongoDB Atlas credentials
username = "admin"
password = "pWXukZL9qbaoLbEG"
cluster_url = "cluster0.ifa9hbf.mongodb.net"
database_name = "ep-db"

def db_connection():
    try:
        # Create a MongoClient instance
        # client = MongoClient(f"mongodb+srv://{username}:{password}@{cluster_url}/?retryWrites=true&w=majority")
        uri = "mongodb+srv://admin:pWXukZL9qbaoLbEG@cluster0.ifa9hbf.mongodb.net/?retryWrites=true&w=majority"
        # Create a new client and connect to the server
        client = MongoClient(uri)
        # Send a ping to confirm a successful connection
        try:
            client.admin.command('ping')
            print("Pinged your deployment. You successfully connected to MongoDB!")
        except Exception as e:
            print("__________________________________________")
            print(e)
            print("__________________________________________")

        print("Connection to MongoDB is successful.")
        return client
    except ConnectionFailure as e:
        print(f"Error connecting to MongoDB: {e}")

def save_property(property):
    # print("Property save: ", property)
    try:
        db = db_connection()
        # property_data = {
        #     "climate_zone": property['climate_zone'],
        #     "full_street_address": property['full_street_address'],
        #     "unit": property['unit'],
        #     "apn": property['apn'],
        #     "owner": property['owner'],
        #     "year_built": property['year_built'],
        #     "square_feet": property['square_feet'],
        #     "lot_size": property['lot_size'],
        #     "bedrooms": property['bedrooms'],
        #     "total_rooms": property['total_rooms'],
        #     "project_extent": property['project_extent'],
        #     "construction_worker": property['construction_worker']
        # }
        property_data = {
            'x': 'x',
            'y': 'y'
        }
        epproperty = db['property']
        print("epproperty: ", epproperty)
        result = epproperty.insert_one(property_data)
        print("result: ", result)
        # result = db.items.insert_one(property_data)
        response = {
            'data': {
                "id": str(result.inserted_id),
                'message': "Property Date save",
                'success': True
            },
        }
        return response
    except ConnectionFailure as e:
        print("Error: ", e)
        response = {
            'data': {
                "id": 'None',
                'message': "Property Date save Error",
                'success': False
            },
        }
        return response

    
